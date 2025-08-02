import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def update_excel_template():
    # Tạo workbook mới
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quiz Template"
    
    # Định nghĩa headers
    headers = [
        "STT", "Câu hỏi", "Đáp án A", "Đáp án B", 
        "Đáp án C", "Đáp án D", "Đáp án đúng", "Thời gian (giây)"
    ]
    
    # Style cho header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Thêm headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Thêm dữ liệu mẫu
    sample_data = [
        [1, "Thủ đô của Việt Nam là gì?", "Hà Nội", "TP.HCM", "Đà Nẵng", "Huế", "A", 30],
        [2, "1 + 1 = ?", "1", "2", "3", "4", "B", 20],
        [3, "Màu của lá cây thường là gì?", "Đỏ", "Vàng", "Xanh", "Trắng", "C", 25],
        [4, "Con vật nào có 4 chân?", "Cá", "Chim", "Chó", "Rắn", "C", 15],
        [5, "Nước nào lớn nhất thế giới?", "Trung Quốc", "Mỹ", "Nga", "Canada", "C", 45]
    ]
    
    # Thêm dữ liệu mẫu
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Điều chỉnh độ rộng cột
    column_widths = [5, 40, 15, 15, 15, 15, 12, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Thêm border cho toàn bộ bảng
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(1, len(sample_data) + 2):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
    
    # Lưu file
    output_path = "DATN_Nhom7/frontend/public/templates/quiz-template-updated.xlsx"
    wb.save(output_path)
    print(f"✅ Đã tạo template mới: {output_path}")
    
    # Tạo file .xls cũng
    try:
        import xlwt
        wb_xls = xlwt.Workbook()
        ws_xls = wb_xls.add_sheet("Quiz Template")
        
        # Thêm headers
        for col, header in enumerate(headers):
            ws_xls.write(0, col, header)
        
        # Thêm dữ liệu mẫu
        for row_idx, row_data in enumerate(sample_data):
            for col_idx, value in enumerate(row_data):
                ws_xls.write(row_idx + 1, col_idx, value)
        
        output_path_xls = "DATN_Nhom7/frontend/public/templates/quiz-template-updated.xls"
        wb_xls.save(output_path_xls)
        print(f"✅ Đã tạo template .xls: {output_path_xls}")
    except ImportError:
        print("⚠️ Không thể tạo file .xls (cần cài xlwt)")

if __name__ == "__main__":
    update_excel_template() 